from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os

def criar_excel_frequencia(materias):
    """Cria um arquivo Excel com os dados de frequência"""
    try:
        # Cria um novo workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Frequência"
        
        # Define os cabeçalhos
        headers = ["Matéria", "Carga Horária", "Faltas", "Frequência", "Máximo de Faltas", "Faltas Restantes", "Status"]
        
        # Estilo para cabeçalhos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center")
        
        # Estilo para células
        cell_alignment = Alignment(horizontal="center")
        
        # Estilo para status
        aprovado_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        reprovado_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Adiciona os cabeçalhos
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Adiciona os dados
        for row, materia in enumerate(materias, 2):
            try:
                # Converte os valores para números
                carga_horaria = float(materia['carga_horaria'].replace(',', '.'))
                faltas = float(materia['faltas'].replace(',', '.'))
                frequencia = float(materia['frequencia'].replace('%', '').replace(',', '.'))
                
                # Calcula o máximo de faltas permitidas (25% da carga horária)
                max_faltas = carga_horaria * 0.25
                
                # Calcula quantas faltas ainda são permitidas
                faltas_restantes = max_faltas - faltas
                
                # Determina o status
                if faltas > max_faltas:
                    status = "REPROVADO POR FALTA"
                else:
                    status = "APROVADO POR FREQUÊNCIA"
                
                # Adiciona os dados na linha
                ws.cell(row=row, column=1, value=materia['nome'])
                ws.cell(row=row, column=2, value=carga_horaria)
                ws.cell(row=row, column=3, value=faltas)
                ws.cell(row=row, column=4, value=frequencia)
                ws.cell(row=row, column=5, value=max_faltas)
                ws.cell(row=row, column=6, value=faltas_restantes)
                ws.cell(row=row, column=7, value=status)
                
                # Aplica o estilo de preenchimento baseado no status
                status_cell = ws.cell(row=row, column=7)
                if status == "APROVADO POR FREQUÊNCIA":
                    status_cell.fill = aprovado_fill
                else:
                    status_cell.fill = reprovado_fill
                
                # Aplica alinhamento central para todas as células da linha
                for col in range(1, 8):
                    ws.cell(row=row, column=col).alignment = cell_alignment
                
            except Exception as e:
                print(f"Erro ao processar matéria {materia['nome']}: {str(e)}")
                continue
        
        # Ajusta a largura das colunas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # Cria a pasta 'relatorios' se não existir
        if not os.path.exists('relatorios'):
            os.makedirs('relatorios')
        
        # Gera o nome do arquivo com data e hora
        data_hora = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f"relatorios/frequencia_{data_hora}.xlsx"
        
        # Salva o arquivo
        wb.save(nome_arquivo)
        print(f"\nArquivo Excel criado com sucesso: {nome_arquivo}")
        return True
        
    except Exception as e:
        print(f"Erro ao criar arquivo Excel: {str(e)}")
        return False 